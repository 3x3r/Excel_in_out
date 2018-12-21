import sys
import csv
import os
import openpyxl
path='C:\\Python27\\test1/'
template = openpyxl.load_workbook(path+"test.xlsx")
temp_sheet = template['Лист1']
b=1
a=1
def copyRange(startCol, startRow, endCol, endRow, sheet):
    rangeSelected = []
    for i in range(startRow, endRow + 1, 1):
        rowSelected = []
        for j in range(startCol, endCol + 1, 1):
            rowSelected.append(sheet.cell(row=i, column=j).value)
        rangeSelected.append(rowSelected)
    return rangeSelected

c=1
d=1
def pasteRange(startCol, startRow, endCol, endRow, sheetReceiving, copiedData):
    countRow = 0
    for i in range(startRow, endRow + 1, 1):
        countCol = 0
        for j in range(startCol, endCol + 1, 1):
            sheetReceiving.cell(row=i, column=j).value = copiedData[countRow][countCol]
            #print(copiedData)
            countCol += 1
        countRow += 1
for rootdir, dirs, files in os.walk(path):
    for file in files:
        if((file.split('.')[-1])=='xlsx'):
            csvpath = os.path.join(rootdir,file)
            wb = openpyxl.load_workbook(csvpath)
            sheet = wb['Лист1']
            selectrange=copyRange(1,14,1,14,sheet)
            pasting=pasteRange(1,b,1,d,temp_sheet,selectrange)

            selectrange=[]
            selectrange = copyRange(1, 21, 7, 21, sheet)
            sheet = wb['Лист1']
            pasting = pasteRange(1, a+1, 7, c+1, temp_sheet, selectrange)
            selectrange = []
            pasting = []
            selectrange = copyRange(1, 22, 7, 22, sheet)
            sheet = wb['Лист1']
            pasting = pasteRange(1, a+2, 7, c+2, temp_sheet, selectrange)
            selectrange = []
            pasting = []
            selectrange = copyRange(1, 23, 7, 23, sheet)
            sheet = wb['Лист1']
            pasting = pasteRange(1, a+3, 7, c+3, temp_sheet, selectrange)
            selectrange = []
            pasting = []
            selectrange = copyRange(1, 24, 7, 24, sheet)
            sheet = wb['Лист1']
            pasting = pasteRange(1, a+4, 7, c+4, temp_sheet, selectrange)
            selectrange = []
            pasting = []
            selectrange = copyRange(1, 25, 7, 25, sheet)
            sheet = wb['Лист1']
            pasting = pasteRange(1, a+5, 7, c+5, temp_sheet, selectrange)
            selectrange = []
            pasting = []
            selectrange = copyRange(1, 26, 7, 26, sheet)
            sheet = wb['Лист1']
            pasting = pasteRange(1, a+6, 7, c+6, temp_sheet, selectrange)
            b=b+1
            d=d+1
            a = a + 10
            b = b + 9
            c = c + 10
            d = d + 9
            selectrange = []
            pasting=[]
            template.save("C:\\Python27\\test1/test.xlsx")
print('End')